"""
Excel workbook export for the Tarantula Market Tracker.
Produces a fully formatted .xlsx with 9 worksheets.
"""
import re
from datetime import datetime
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from models import Listing, CrawlResult, DealRating

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
C_HEADER_BG   = "1F3864"   # Dark navy header
C_HEADER_FG   = "FFFFFF"
C_DEAL_EXC    = "FFD700"   # Gold  💎💎
C_DEAL_STRONG = "C6EFCE"   # Green 💎
C_DEAL_FAIR   = "FFFFFF"   # White 👍
C_DEAL_ABOVE  = "FFDCE1"   # Pink  👎
C_ALT_ROW     = "F0F4F8"   # Light blue-grey alt row
C_FROZEN_HDR  = "2E6D8F"   # Mid-blue for frozen header rows


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _header_font() -> Font:
    return Font(bold=True, color=C_HEADER_FG, size=10, name="Calibri")


def _border() -> Border:
    thin = Side(style="thin", color="D0D7DE")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _deal_fill(rating: Optional[str]) -> Optional[PatternFill]:
    if rating == "🔥":
        return PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
    if rating == DealRating.EXCEPTIONAL:
        return _fill(C_DEAL_EXC)
    if rating == DealRating.STRONG:
        return _fill(C_DEAL_STRONG)
    if rating == DealRating.ABOVE_MARKET:
        return _fill(C_DEAL_ABOVE)
    return None


def _avail_display(avail: str) -> str:
    mapping = {
        "in_stock": "In Stock",
        "out_of_stock": "Out of Stock",
        "preorder": "Preorder",
        "limited": "Limited",
        "unknown": "Unknown",
    }
    return mapping.get(avail, avail)


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------
def export_workbook(
    listings: list[Listing],
    crawl_results: list[CrawlResult],
    output_path: Path,
) -> None:
    """Create or update the master Excel workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    ws_all      = wb.create_sheet("All Listings")
    ws_deals    = wb.create_sheet("Best Deals")
    ws_females  = wb.create_sheet("Females")
    ws_unsexed  = wb.create_sheet("Unsexed & Juveniles")
    ws_vendors  = wb.create_sheet("Vendor Summary")
    ws_species  = wb.create_sheet("Species Summary")
    ws_history  = wb.create_sheet("Price History")
    ws_status   = wb.create_sheet("Crawl Status")
    ws_method   = wb.create_sheet("Methodology")

    # Build sheets
    _build_all_listings(ws_all, listings)
    _build_best_deals(ws_deals, listings)
    _build_females(ws_females, listings)
    _build_unsexed(ws_unsexed, listings)
    _build_vendor_summary(ws_vendors, listings)
    _build_species_summary(ws_species, listings)
    _build_price_history(ws_history, listings)
    _build_crawl_status(ws_status, crawl_results)
    _build_methodology(ws_method)

    wb.save(str(output_path))
    print(f"[EXPORT] Workbook saved: {output_path}")


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _apply_header(ws, headers: list[str], col_widths: dict = None):
    """Write a styled header row."""
    ws.row_dimensions[1].height = 20
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _header_font()
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()

    # Set column widths
    if col_widths:
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_row(ws, row_num: int, values: list, rating: Optional[str] = None, alt: bool = False):
    """Write a data row with optional deal coloring."""
    fill = _deal_fill(rating)
    if fill is None and alt:
        fill = _fill(C_ALT_ROW)

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = _font()
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = _border()
        if fill:
            cell.fill = fill

        # Format currency columns
        if isinstance(val, float) and val > 0 and col >= 6:
            cell.number_format = '"$"#,##0.00'


def _build_all_listings(ws, listings: list[Listing]):
    headers = [
        "Deal", "Scientific Name", "Common Name", "Vendor", "Sex", "Size",
        "Price", "Regular Price", "Discount %", "Availability", "Qty",
        "Current Lowest", "Market Avg", "Hist. Low", "Price/Inch",
        "Date Checked", "Product URL", "Notes",
    ]
    widths = {
        1: 6, 2: 32, 3: 22, 4: 22, 5: 14, 6: 10,
        7: 10, 8: 12, 9: 10, 10: 13, 11: 6,
        12: 13, 13: 11, 14: 10, 15: 10,
        16: 12, 17: 40, 18: 35,
    }
    _apply_header(ws, headers, widths)

    from models import DealRating
    DEAL_SORT = {DealRating.EXCEPTIONAL: 0, DealRating.STRONG: 1,
                 DealRating.FAIR: 2, DealRating.ABOVE_MARKET: 3, None: 4}
    sorted_listings = sorted(listings, key=lambda l: (
        DEAL_SORT.get(l.deal_rating, 4), l.scientific_name_key or l.scientific_name
    ))

    for i, l in enumerate(sorted_listings, 2):
        disc = l.discount_pct()
        row = [
            l.deal_rating or "",
            l.scientific_name,
            l.common_name or "",
            l.vendor,
            l.sex_display,
            l.size_text or "",
            l.price_usd,
            l.regular_price_usd or "",
            f"{disc:.0f}%" if disc else "",
            _avail_display(l.availability),
            l.quantity or "",
            l.current_lowest_price or "",
            l.market_average or "",
            l.historical_low or "",
            l.price_per_inch or "",
            l.date_checked.strftime("%Y-%m-%d") if l.date_checked else "",
            l.product_url,
            _build_notes(l),
        ]
        _write_row(ws, i, row, l.deal_rating, alt=(i % 2 == 0))

    # Freeze and bold first column
    ws.freeze_panes = "B2"


def _build_best_deals(ws, listings: list[Listing]):
    from models import DealRating, Availability
    SORT_ORDER = {DealRating.EXCEPTIONAL: 0, DealRating.STRONG: 1,
                  DealRating.FAIR: 2, DealRating.ABOVE_MARKET: 3, None: 4}

    deals = [l for l in listings if l.availability != Availability.OUT_OF_STOCK]
    deals.sort(key=lambda l: (
        SORT_ORDER.get(l.deal_rating, 4),
        -(l.market_average or 0) + (l.price_usd or 0),  # biggest pct below market first
        0 if l.sex == "F" else 1,
        l.price_usd or 9999,
    ))

    headers = [
        "Deal", "Scientific Name", "Common Name", "Vendor", "Sex", "Size",
        "Price", "Market Avg", "% Below Mkt", "Hist. Low", "Availability", "URL", "Deal Reason",
    ]
    widths = {1: 6, 2: 32, 3: 22, 4: 22, 5: 14, 6: 10,
              7: 10, 8: 11, 9: 11, 10: 10, 11: 13, 12: 40, 13: 40}
    _apply_header(ws, headers, widths)

    for i, l in enumerate(deals, 2):
        pct_below = ""
        if l.market_average and l.price_usd:
            d = (l.market_average - l.price_usd) / l.market_average * 100
            pct_below = f"{d:.0f}%"
        row = [
            l.deal_rating or "",
            l.scientific_name,
            l.common_name or "",
            l.vendor,
            l.sex_display,
            l.size_text or "",
            l.price_usd,
            l.market_average or "",
            pct_below,
            l.historical_low or "",
            _avail_display(l.availability),
            l.product_url,
            l.deal_reason or "",
        ]
        _write_row(ws, i, row, l.deal_rating, alt=(i % 2 == 0))


def _build_females(ws, listings: list[Listing]):
    females = sorted(
        [l for l in listings if l.sex in ("F", "PF")],
        key=lambda l: (l.scientific_name_key or l.scientific_name, l.price_usd)
    )
    headers = [
        "Deal", "Scientific Name", "Common Name", "Vendor", "Sex", "Size",
        "Price", "Current Lowest", "Market Avg", "Hist. Low", "Date", "URL", "Notes",
    ]
    widths = {1: 6, 2: 32, 3: 22, 4: 22, 5: 14, 6: 10,
              7: 10, 8: 13, 9: 11, 10: 10, 11: 12, 12: 40, 13: 30}
    _apply_header(ws, headers, widths)

    for i, l in enumerate(females, 2):
        row = [
            l.deal_rating or "",
            l.scientific_name,
            l.common_name or "",
            l.vendor,
            l.sex_display,
            l.size_text or "",
            l.price_usd,
            l.current_lowest_price or "",
            l.market_average or "",
            l.historical_low or "",
            l.date_checked.strftime("%Y-%m-%d") if l.date_checked else "",
            l.product_url,
            _build_notes(l),
        ]
        _write_row(ws, i, row, l.deal_rating, alt=(i % 2 == 0))


def _build_unsexed(ws, listings: list[Listing]):
    unsexed = sorted(
        [l for l in listings if l.sex in ("U", "Unknown")],
        key=lambda l: (l.scientific_name_key or l.scientific_name, l.price_usd)
    )
    headers = [
        "Deal", "Scientific Name", "Common Name", "Vendor", "Size",
        "Price", "Market Avg", "Availability", "Date", "URL", "Notes",
    ]
    widths = {1: 6, 2: 32, 3: 22, 4: 22, 5: 10,
              6: 10, 7: 11, 8: 13, 9: 12, 10: 40, 11: 30}
    _apply_header(ws, headers, widths)

    for i, l in enumerate(unsexed, 2):
        row = [
            l.deal_rating or "",
            l.scientific_name,
            l.common_name or "",
            l.vendor,
            l.size_text or "",
            l.price_usd,
            l.market_average or "",
            _avail_display(l.availability),
            l.date_checked.strftime("%Y-%m-%d") if l.date_checked else "",
            l.product_url,
            _build_notes(l),
        ]
        _write_row(ws, i, row, l.deal_rating, alt=(i % 2 == 0))


def _build_vendor_summary(ws, listings: list[Listing]):
    from models import DealRating
    headers = [
        "Vendor", "Active Listings", "Female Listings", "Lowest Price",
        "Avg Price", "💎💎 Count", "💎 Count", "👍 Count", "👎 Count", "Date Checked",
    ]
    widths = {1: 28, 2: 15, 3: 15, 4: 13, 5: 11,
              6: 10, 7: 10, 8: 10, 9: 10, 10: 14}
    _apply_header(ws, headers, widths)

    vendors = {}
    for l in listings:
        v = l.vendor
        if v not in vendors:
            vendors[v] = []
        vendors[v].append(l)

    for i, (vendor, vl) in enumerate(sorted(vendors.items()), 2):
        active = [l for l in vl if l.availability != "out_of_stock"]
        females = [l for l in active if l.sex in ("F", "PF")]
        prices = [l.price_usd for l in active if l.price_usd]
        dates = [l.date_checked for l in vl if l.date_checked]
        row = [
            vendor,
            len(active),
            len(females),
            min(prices) if prices else "",
            round(sum(prices) / len(prices), 2) if prices else "",
            sum(1 for l in active if l.deal_rating == DealRating.EXCEPTIONAL),
            sum(1 for l in active if l.deal_rating == DealRating.STRONG),
            sum(1 for l in active if l.deal_rating == DealRating.FAIR),
            sum(1 for l in active if l.deal_rating == DealRating.ABOVE_MARKET),
            max(dates).strftime("%Y-%m-%d") if dates else "",
        ]
        _write_row(ws, i, row, alt=(i % 2 == 0))


def _build_species_summary(ws, listings: list[Listing]):
    headers = [
        "Scientific Name", "Common Name", "Active Vendors", "Active Listings",
        "Lowest Price", "Lowest Female Price", "Avg Price",
        "Hist. Low", "Best Vendor", "Last Checked",
    ]
    widths = {1: 32, 2: 22, 3: 15, 4: 15, 5: 13, 6: 17,
              7: 11, 8: 10, 9: 22, 10: 14}
    _apply_header(ws, headers, widths)

    species = {}
    for l in listings:
        key = l.scientific_name_key or l.scientific_name.lower()
        if key not in species:
            species[key] = {"name": l.scientific_name, "common": l.common_name, "listings": []}
        species[key]["listings"].append(l)

    for i, (key, data) in enumerate(sorted(species.items()), 2):
        sl = data["listings"]
        active = [l for l in sl if l.availability != "out_of_stock"]
        females = [l for l in active if l.sex in ("F", "PF")]
        prices = [l.price_usd for l in active if l.price_usd]
        f_prices = [l.price_usd for l in females if l.price_usd]
        vendors = list({l.vendor for l in active})
        dates = [l.date_checked for l in sl if l.date_checked]
        best_vendor = min(active, key=lambda l: l.price_usd or 9999).vendor if active else ""

        row = [
            data["name"],
            data["common"] or "",
            len(vendors),
            len(active),
            min(prices) if prices else "",
            min(f_prices) if f_prices else "",
            round(sum(prices) / len(prices), 2) if prices else "",
            min(l.historical_low for l in sl if l.historical_low) if any(l.historical_low for l in sl) else "",
            best_vendor,
            max(dates).strftime("%Y-%m-%d") if dates else "",
        ]
        _write_row(ws, i, row, alt=(i % 2 == 0))


def _build_price_history(ws, listings: list[Listing]):
    headers = [
        "Date", "Vendor", "Scientific Name", "Sex", "Size",
        "Price", "Availability", "URL",
    ]
    widths = {1: 12, 2: 22, 3: 32, 4: 14, 5: 10, 6: 10, 7: 13, 8: 40}
    _apply_header(ws, headers, widths)

    sorted_l = sorted(listings, key=lambda l: (
        l.date_checked or datetime.min,), reverse=True)

    for i, l in enumerate(sorted_l, 2):
        row = [
            l.date_checked.strftime("%Y-%m-%d %H:%M") if l.date_checked else "",
            l.vendor,
            l.scientific_name,
            l.sex_display,
            l.size_text or "",
            l.price_usd,
            _avail_display(l.availability),
            l.product_url,
        ]
        _write_row(ws, i, row, alt=(i % 2 == 0))


def _build_crawl_status(ws, results: list[CrawlResult]):
    headers = [
        "Vendor", "Status", "Pages Crawled", "Products Found",
        "Variants Found", "Failures", "Started", "Finished", "Duration (s)", "Notes",
    ]
    widths = {1: 28, 2: 12, 3: 14, 4: 14, 5: 14,
              6: 10, 7: 18, 8: 18, 9: 12, 10: 40}
    _apply_header(ws, headers, widths)

    for i, r in enumerate(results, 2):
        dur = r.duration_seconds()
        row = [
            r.vendor_name,
            r.status,
            r.pages_crawled,
            r.products_found,
            r.variants_found,
            len(r.failures),
            r.started_at.strftime("%Y-%m-%d %H:%M:%S") if r.started_at else "",
            r.finished_at.strftime("%Y-%m-%d %H:%M:%S") if r.finished_at else "",
            round(dur, 1) if dur else "",
            r.notes or ("; ".join(r.failures[:3]) if r.failures else ""),
        ]
        status_fill = {
            "complete": _fill("C6EFCE"),
            "partial": _fill("FFEB9C"),
            "failed": _fill("FFDCE1"),
            "running": _fill("DEEBF7"),
        }.get(r.status)
        _write_row(ws, i, row, alt=(i % 2 == 0))
        if status_fill:
            for col in range(1, len(row) + 1):
                ws.cell(row=i, column=col).fill = status_fill


def _build_methodology(ws):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80

    sections = [
        ("METHODOLOGY", ""),
        ("", ""),
        ("Deal Ratings", ""),
        ("💎💎 Exceptional Deal", (
            "Used when: (a) price is 20% or more below comparable market median, "
            "(b) price matches or beats the all-time historical low, "
            "(c) confirmed female listed unusually cheaply for size/species, or "
            "(d) it is the cheapest confirmed female in the current dataset."
        )),
        ("💎 Strong Deal", "10-20% below comparable market median, or top quartile value for same species/sex/size."),
        ("👍 Fair Market Price", "Within approximately 10% of comparable market median."),
        ("👎 Above Market", "More than 10-15% above comparable market median without a justifying size, sex, rarity, or locality difference."),
        ("", ""),
        ("Comparison Rules", ""),
        ("Sex grouping", (
            "Female (F) and Probable Female (PF) are never compared against Unsexed (U) or Male (M). "
            "Each sex group is scored independently. Mature Male (MM) is not compared against juvenile males."
        )),
        ("Size grouping", (
            "Listings are bucketed by approximate size: <0.5\" (XS), 0.5-1\" (Sling), 1-2\" (Juvenile), "
            "2-3.5\" (Sub-adult), 3.5-5.5\" (Adult-S), 5.5\"+ (Adult-L). "
            "Only listings in the same bucket are compared."
        )),
        ("'From' prices", (
            "A displayed 'from $X' price is NEVER recorded as the price for a specific sex/size variant. "
            "It is flagged with verification_level=estimated and noted in the Notes column."
        )),
        ("Bulk pricing", (
            "Stored with package_quantity, package_price, and per_animal_price. "
            "Per-animal price is used for deal comparison when bulk_quantity >= 2."
        )),
        ("", ""),
        ("Data Limitations", ""),
        ("Verification levels", (
            "direct = crawled directly from vendor product page. "
            "aggregator = found via TarantulaList or similar listing site. "
            "estimated = inferred from non-specific pricing (e.g., 'from' price)."
        )),
        ("Historical lows", (
            "Historical lows are computed from all price_history observations in the SQLite database. "
            "On the first crawl, historical_low equals the current lowest observed price."
        )),
        ("Crawl date", f"Methodology sheet last updated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"),
    ]

    row = 1
    for label, text in sections:
        if label == "METHODOLOGY":
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = Font(bold=True, size=14, color="1F3864", name="Calibri")
        elif label == "" and text == "":
            row += 1
            continue
        elif text == "":
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = Font(bold=True, size=11, color="2E6D8F", name="Calibri")
            ws.cell(row=row, column=1).fill = _fill("DEEBF7")
        else:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10, name="Calibri")
            text_cell = ws.cell(row=row, column=2, value=text)
            text_cell.font = Font(size=10, name="Calibri")
            text_cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = max(30, len(text) // 6)
        row += 1


def _build_notes(l: Listing) -> str:
    """Build the Notes column value."""
    parts = []
    if l.notes:
        parts.append(l.notes)
    if l.bulk_quantity:
        parts.append(f"Bulk: {l.bulk_quantity} for ${l.bulk_package_price:.2f} (${l.bulk_per_animal_price:.2f}/ea)")
    if l.is_new:
        parts.append("NEW LISTING")
    if l.is_price_drop:
        prev = l.previous_price
        parts.append(f"PRICE DROP from ${prev:.2f}" if prev else "PRICE DROP")
    if l.is_new_historical_low:
        parts.append("NEW HISTORICAL LOW")
    if l.is_returned_to_stock:
        parts.append("BACK IN STOCK")
    if l.deal_reason:
        parts.append(l.deal_reason)
    return " | ".join(parts)


# ──────────────────────────────────────────────────────────────────────────────
# build_workbook — dict-friendly entry point called by pipeline.py
# ──────────────────────────────────────────────────────────────────────────────

def build_workbook(
    snapshot:          list,         # list of price_history row dicts (current listings)
    full_history:      list,         # list of ALL price_history row dicts
    crawl_summary:     list,         # list of crawl_run row dicts from get_crawl_summary()
    output_path:       str,
    rarity_data:       dict = None,  # {species_key: RarityData} from compute_all_rarity()
    size_class_rarity: dict = None,  # {(species_key, size_bucket): SizeClassRarity}
    dest_zip:          str  = "00000",
) -> None:
    """
    Pipeline entry point. Accepts dict-format data from the history engine
    and builds the full 9-sheet workbook, including the live Price History
    and Crawl Status sheets populated from the DB.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from pathlib import Path as _Path

    out = _Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_all      = wb.create_sheet("All Listings")
    ws_deals    = wb.create_sheet("Best Deals")
    ws_females  = wb.create_sheet("Females")
    ws_unsexed  = wb.create_sheet("Unsexed & Juveniles")
    ws_vendors  = wb.create_sheet("Vendor Summary")
    ws_species  = wb.create_sheet("Species Summary")
    ws_history  = wb.create_sheet("Price History")
    ws_status   = wb.create_sheet("Crawl Status")
    ws_method   = wb.create_sheet("Methodology")

    # Inject rarity data into snapshot dicts
    if rarity_data:
        for l in snapshot:
            key = l.get("scientific_name_key", "")
            rd = rarity_data.get(key)
            if rd:
                l["rarity_score"]          = rd.score
                l["rarity_label"]          = rd.label
                l["new_to_system"]         = rd.new_to_system
                l["rarity_vendor_count"]   = rd.vendor_count
                l["rarity_current_sellers"]= len(rd.current_vendors)
                l["last_seen_date"]        = rd.last_seen
                l["last_seen_price"]       = rd.last_price

    # Inject size-class rarity data
    if size_class_rarity:
        from scoring.deals import _size_bucket
        from scoring.rarity import SIZE_BUCKET_DISPLAY
        for l in snapshot:
            key = l.get("scientific_name_key", "")
            mid = l.get("size_midpoint")
            sb  = _size_bucket(mid)
            scr = size_class_rarity.get((key, sb))
            l["size_bucket"]              = sb
            l["size_bucket_label"]        = SIZE_BUCKET_DISPLAY.get(sb, sb)
            if scr:
                l["size_class_rarity_score"]  = scr.score
                l["size_class_rarity_label"]  = scr.label
                l["size_class_vendor_count"]  = scr.vendor_count
                l["size_class_obs_count"]     = scr.obs_count
                l["size_class_last_seen"]     = scr.last_seen
                l["size_class_last_price"]    = scr.last_price

    _build_all_listings_dict(ws_all,    snapshot)
    _build_best_deals_dict(ws_deals,    snapshot)
    _build_females_dict(ws_females,     snapshot)
    _build_unsexed_dict(ws_unsexed,     snapshot)
    _build_vendor_summary_dict(ws_vendors, snapshot)
    _build_species_summary_dict(ws_species, snapshot)
    if rarity_data or size_class_rarity:
        ws_rarity = wb.create_sheet("Rarity Index", 6)
        _build_rarity_index_dict(ws_rarity, rarity_data or {}, size_class_rarity or {})
    _build_price_history_dict(ws_history, full_history)
    _build_crawl_status_dict(ws_status, crawl_summary)
    _build_methodology(ws_method)

    wb.save(str(out))
    n_sheets = len(wb.sheetnames)
    print(f"[EXPORT] Workbook saved: {out} ({len(snapshot)} active listings, "
          f"{len(full_history)} history rows, {len(crawl_summary)} crawl runs, "
          f"{n_sheets} sheets)")


# ── Dict-based sheet builders ─────────────────────────────────────────────────

def _g(row: dict, *keys, default=""):
    """Safe dict getter that tries multiple key names."""
    for k in keys:
        v = row.get(k)
        if v is not None and v != "":
            return v
    return default


def _build_all_listings_dict(ws, rows: list) -> None:
    headers = ["Vendor", "Scientific Name", "Common Name", "Sex", "Size",
               "Price", "Landed", "Ship", "Deal", "🔥",
               "Rarity", "Score", "Sellers Now",
               "Hist Low", "Mkt Avg", "% vs Median",
               "Last Seen", "Last Price", "Verification", "Observed"]
    widths  = {1:22, 2:32, 3:22, 4:8, 5:8, 6:8, 7:9, 8:7, 9:6, 10:4,
               11:22, 12:6, 13:12,
               14:9, 15:9, 16:12,
               17:12, 18:10, 19:14, 20:12}
    _apply_header(ws, headers, widths)

    for i, r in enumerate(rows, 2):
        price = _g(r, "price_usd", "price") or 0
        avg   = _g(r, "market_average") or 0
        pct   = round((price - avg) / avg * 100, 1) if avg else ""
        landed   = _g(r, "landed_cost")
        shipping = _g(r, "shipping_share")
        is_fire  = _g(r, "is_fire_deal")
        deal_r   = _g(r, "deal_rating")
        rarity_s = _g(r, "rarity_score")
        rarity_l = _g(r, "rarity_label")
        row   = [
            _g(r, "vendor_key"),
            _g(r, "scientific_name"),
            _g(r, "common_name"),
            _g(r, "sex_display", "sex"),
            _g(r, "size_text"),
            price,
            round(landed, 2) if landed else "",
            round(shipping, 2) if shipping else "",
            deal_r,
            "🔥" if is_fire else "",
            rarity_l or "",
            rarity_s or "",
            _g(r, "rarity_current_sellers"),
            _g(r, "historical_low"),
            avg or "",
            pct,
            _g(r, "last_seen_date"),
            f"${_g(r, 'last_seen_price'):.2f}" if _g(r, "last_seen_price") else "",
            _g(r, "verification_level"),
            str(_g(r, "observed_at"))[:10],
        ]
        # 🔥 rows get gold highlight
        _write_row(ws, i, row, rating=deal_r if not is_fire else "🔥", alt=(i % 2 == 0))
        if is_fire:
            for col in range(1, 10):
                ws.cell(row=i, column=col).fill = _fill("FF8F00")  # deep amber
                ws.cell(row=i, column=col).font = Font(bold=True, color="FFFFFF", size=10)


def _build_best_deals_dict(ws, rows: list) -> None:
    deals = [r for r in rows if _g(r, "deal_rating") in ("💎💎", "💎")]
    deals.sort(key=lambda r: (_g(r, "deal_rating") != "💎💎", _g(r, "price_usd", "price") or 999))
    headers = ["Deal", "Vendor", "Scientific Name", "Sex", "Size", "Price",
               "Hist Low", "Mkt Avg", "Why", "Observed"]
    widths  = {1:5, 2:22, 3:32, 4:8, 5:8, 6:8, 7:9, 8:9, 9:50, 10:12}
    _apply_header(ws, headers, widths)
    for i, r in enumerate(deals, 2):
        price = _g(r, "price_usd", "price") or 0
        row = [
            _g(r, "deal_rating"),
            _g(r, "vendor_key"),
            _g(r, "scientific_name"),
            _g(r, "sex_display", "sex"),
            _g(r, "size_text"),
            price,
            _g(r, "historical_low"),
            _g(r, "market_average"),
            _g(r, "deal_reason"),
            str(_g(r, "observed_at"))[:10],
        ]
        _write_row(ws, i, row, rating=_g(r, "deal_rating"), alt=(i % 2 == 0))


def _build_females_dict(ws, rows: list) -> None:
    females = [r for r in rows if _g(r, "sex") in ("F", "PF")]
    females.sort(key=lambda r: (_g(r, "scientific_name"), _g(r, "price_usd", "price") or 999))
    headers = ["Vendor", "Scientific Name", "Sex", "Size", "Price", "Deal", "Hist Low", "Observed"]
    widths  = {1:22, 2:35, 3:8, 4:8, 5:8, 6:6, 7:9, 8:12}
    _apply_header(ws, headers, widths)
    for i, r in enumerate(females, 2):
        row = [
            _g(r, "vendor_key"),
            _g(r, "scientific_name"),
            _g(r, "sex_display", "sex"),
            _g(r, "size_text"),
            _g(r, "price_usd", "price"),
            _g(r, "deal_rating"),
            _g(r, "historical_low"),
            str(_g(r, "observed_at"))[:10],
        ]
        _write_row(ws, i, row, rating=_g(r, "deal_rating"), alt=(i % 2 == 0))


def _build_unsexed_dict(ws, rows: list) -> None:
    unsexed = [r for r in rows if _g(r, "sex") in ("U", "Unknown", None, "")]
    unsexed.sort(key=lambda r: (_g(r, "scientific_name"), _g(r, "price_usd", "price") or 999))
    headers = ["Vendor", "Scientific Name", "Size", "Price", "Deal", "Mkt Avg", "Observed"]
    widths  = {1:22, 2:35, 3:8, 4:8, 5:6, 6:9, 7:12}
    _apply_header(ws, headers, widths)
    for i, r in enumerate(unsexed, 2):
        row = [
            _g(r, "vendor_key"),
            _g(r, "scientific_name"),
            _g(r, "size_text"),
            _g(r, "price_usd", "price"),
            _g(r, "deal_rating"),
            _g(r, "market_average"),
            str(_g(r, "observed_at"))[:10],
        ]
        _write_row(ws, i, row, rating=_g(r, "deal_rating"), alt=(i % 2 == 0))


def _build_vendor_summary_dict(ws, rows: list) -> None:
    from collections import Counter
    counts = Counter(_g(r, "vendor_key") for r in rows)
    headers = ["Vendor", "Listing Count", "Avg Price", "Min Price", "Max Price",
               "Deals (💎+)"]
    widths  = {1:28, 2:14, 3:11, 4:11, 5:11, 6:12}
    _apply_header(ws, headers, widths)
    for i, (vk, cnt) in enumerate(sorted(counts.items()), 2):
        vrows = [r for r in rows if _g(r, "vendor_key") == vk]
        prices = [_g(r, "price_usd", "price") or 0 for r in vrows if (_g(r, "price_usd", "price") or 0) > 0]
        deals  = sum(1 for r in vrows if _g(r, "deal_rating") in ("💎💎", "💎"))
        row = [
            vk, cnt,
            round(sum(prices)/len(prices), 2) if prices else "",
            min(prices) if prices else "",
            max(prices) if prices else "",
            deals,
        ]
        _write_row(ws, i, row, alt=(i % 2 == 0))


def _build_species_summary_dict(ws, rows: list) -> None:
    from collections import defaultdict
    species = defaultdict(list)
    for r in rows:
        species[_g(r, "scientific_name")].append(r)
    headers = ["Scientific Name", "Common Name", "Listings", "Min Price", "Max Price",
               "Median", "Best Deal", "Vendors"]
    widths  = {1:35, 2:25, 3:9, 4:10, 5:10, 6:9, 7:9, 8:30}
    _apply_header(ws, headers, widths)
    rows_out = []
    for sci, srows in species.items():
        prices = [_g(r, "price_usd", "price") or 0 for r in srows if (_g(r, "price_usd", "price") or 0) > 0]
        if not prices: continue
        import statistics as _st
        best = min((_g(r, "deal_rating") for r in srows if _g(r, "deal_rating")), default="")
        vk_set = {_g(r, "vendor_key") for r in srows}
        rows_out.append((sci, srows[0], prices, best, vk_set))
    rows_out.sort(key=lambda x: x[0])
    for i, (sci, srow, prices, best, vk_set) in enumerate(rows_out, 2):
        import statistics as _st
        row = [
            sci,
            _g(srow, "common_name"),
            len(prices),
            min(prices),
            max(prices),
            round(_st.median(prices), 2),
            best,
            ", ".join(sorted(vk_set))[:100],
        ]
        _write_row(ws, i, row, rating=best, alt=(i % 2 == 0))


def _build_price_history_dict(ws, history_rows: list) -> None:
    """Populate Price History sheet from all DB history rows."""
    headers = [
        "Date", "Vendor", "Source Type", "Scientific Name", "Common Name",
        "Sex", "Size", "Price", "Was Price",
        "Deal", "↓Drop", "★Low", "Notes", "Run ID",
    ]
    widths = {1:12, 2:22, 3:16, 4:32, 5:22, 6:8, 7:8, 8:8, 9:8, 10:6, 11:6, 12:6, 13:35, 14:7}
    _apply_header(ws, headers, widths)

    for i, r in enumerate(history_rows, 2):
        price_drop = _g(r, "is_price_drop")
        new_low    = _g(r, "is_new_historical_low")
        row = [
            str(_g(r, "observed_at"))[:10],
            _g(r, "vendor_key"),
            _g(r, "platform", "verification_level"),
            _g(r, "scientific_name"),
            _g(r, "common_name"),
            _g(r, "sex_display", "sex"),
            _g(r, "size_text"),
            _g(r, "price_usd"),
            _g(r, "regular_price_usd"),
            _g(r, "deal_rating"),
            "↓" if price_drop else "",
            "★" if new_low    else "",
            _g(r, "notes"),
            _g(r, "crawl_run_id"),
        ]
        _write_row(ws, i, row, rating=_g(r, "deal_rating"), alt=(i % 2 == 0))

        # Highlight price drops in light orange
        if price_drop and not new_low:
            for col in range(1, 9):
                ws.cell(row=i, column=col).fill = _fill("FFE0B2")
        # Highlight new historical lows in gold
        if new_low:
            for col in range(1, 9):
                ws.cell(row=i, column=col).fill = _fill("FFF176")


def _build_crawl_status_dict(ws, crawl_rows: list) -> None:
    """Populate Crawl Status sheet from crawl_summary dicts."""
    headers = [
        "Run ID", "Vendor", "Platform", "Status", "Listings",
        "Min Price", "Max Price", "Started", "Notes",
    ]
    widths = {1:7, 2:28, 3:16, 4:10, 5:10, 6:10, 7:10, 8:18, 9:50}
    _apply_header(ws, headers, widths)

    STATUS_COLORS = {
        "complete": "C6EFCE",
        "partial":  "FFEB9C",
        "failed":   "FFDCE1",
        "running":  "DEEBF7",
    }

    for i, r in enumerate(crawl_rows, 2):
        status = _g(r, "status")
        row = [
            _g(r, "id"),
            _g(r, "vendor_name", "vendor_key"),
            _g(r, "platform"),
            status,
            _g(r, "ph_rows", "variants_found"),
            _g(r, "min_price"),
            _g(r, "max_price"),
            str(_g(r, "started_at"))[:16],
            _g(r, "notes"),
        ]
        _write_row(ws, i, row, alt=(i % 2 == 0))
        fill_hex = STATUS_COLORS.get(status)
        if fill_hex:
            for col in range(1, len(row) + 1):
                ws.cell(row=i, column=col).fill = _fill(fill_hex)


def _build_rarity_index_sheet(ws, rarity_data: dict) -> None:
    """
    Rarity Index sheet: every known species ranked by rarity score.
    Columns: Score | Label | Species | Vendors (ever) | Obs | Currently Selling |
             First Seen | Last Seen | Last Price | Last Vendor | New?
    """
    from scoring.rarity import RARITY_LABELS

    headers = [
        "Score", "Rarity", "Scientific Name",
        "Vendors (ever)", "Observations", "Currently Selling",
        "First Seen", "Last Seen", "Last Price", "Last Vendor", "🆕 New?"
    ]
    widths = {1:7, 2:22, 3:38, 4:14, 5:12, 6:17,
              7:12, 8:12, 9:11, 10:22, 11:8}
    _apply_header(ws, headers, widths)

    # Score → fill color
    SCORE_FILLS = {
        10: "7B1FA2",  # deep purple — scientific specimen
        9:  "4A148C",  # purple
        8:  "6A1B9A",  # medium purple
        7:  "1565C0",  # dark blue
        6:  "1976D2",  # blue
        5:  "2E7D32",  # dark green
        4:  "558B2F",  # olive green
        3:  "F57F17",  # dark amber
        2:  "757575",  # grey
        1:  "424242",  # dark grey
    }
    TEXT_WHITE = Font(bold=True, color="FFFFFF", size=10)
    TEXT_DARK  = Font(color="212121", size=10)

    # Sort: highest rarity first, then by species name
    sorted_species = sorted(rarity_data.values(),
                            key=lambda rd: (-rd.score, rd.species_key))

    for i, rd in enumerate(sorted_species, 2):
        new_flag  = "🆕" if rd.new_to_system else ""
        row = [
            rd.score,
            rd.label,
            rd.species_key,
            rd.vendor_count,
            rd.obs_count,
            len(rd.current_vendors),
            rd.first_seen or "",
            rd.last_seen  or "",
            f"${rd.last_price:.2f}" if rd.last_price else "",
            rd.last_vendor or "",
            new_flag,
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = _border()
            cell.alignment = Alignment(vertical="center")

        # Color-code score column
        score_hex = SCORE_FILLS.get(rd.score, "FFFFFF")
        ws.cell(row=i, column=1).fill = PatternFill(
            start_color=score_hex, end_color=score_hex, fill_type="solid"
        )
        ws.cell(row=i, column=1).font = TEXT_WHITE
        ws.cell(row=i, column=2).font = Font(
            bold=(rd.score >= 7), color=score_hex, size=10
        )
        if rd.new_to_system:
            for col in range(1, 12):
                if col not in (1, 2):
                    ws.cell(row=i, column=col).fill = _fill("FFF9C4")


def add_rarity_index_sheet(wb, rarity_data: dict) -> None:
    """Insert or update the Rarity Index worksheet in an existing workbook."""
    # Remove existing sheet if present
    if "Rarity Index" in wb.sheetnames:
        del wb["Rarity Index"]
    ws = wb.create_sheet("Rarity Index", 6)  # after Species Summary
    _build_rarity_index_sheet(ws, rarity_data)


def _build_rarity_index_dict(ws, rarity_data: dict, size_class_rarity: dict) -> None:
    """
    Rarity Index sheet with two views:
    - Top section: species-level rarity ranked highest to lowest
    - Per row: drills into size-class breakdown for that species

    Columns:
    Spc Rarity | Size Class | Sz Rarity | Species | Life Stage |
    Vendors(ever) | Obs | Sellers Now | Last Seen | Last Price | New?
    """
    from scoring.rarity import RARITY_LABELS, SIZE_BUCKET_DISPLAY

    headers = [
        "Spc\nScore", "Species Rarity", "Scientific Name",
        "Life Stage", "Sz\nScore", "Size Class Rarity",
        "Vendors\n(ever)", "Obs", "Sellers\nNow",
        "Last Seen", "Last Price", "🆕"
    ]
    widths = {1:7, 2:22, 3:38, 4:16, 5:7, 6:22,
              7:9, 8:6, 9:9, 10:12, 11:11, 12:5}
    _apply_header(ws, headers, widths)

    SCORE_COLORS = {
        10: ("7B1FA2", "FFFFFF"),  # deep purple / white
        9:  ("4527A0", "FFFFFF"),
        8:  ("1565C0", "FFFFFF"),
        7:  ("1976D2", "FFFFFF"),
        6:  ("0277BD", "FFFFFF"),
        5:  ("2E7D32", "FFFFFF"),
        4:  ("558B2F", "FFFFFF"),
        3:  ("F57F17", "000000"),
        2:  ("757575", "FFFFFF"),
        1:  ("424242", "FFFFFF"),
    }

    def score_cell(ws, row_i, col, score, text):
        cell = ws.cell(row=row_i, column=col, value=text)
        if score:
            bg, fg = SCORE_COLORS.get(score, ("FFFFFF", "000000"))
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            cell.font = Font(bold=True, color=fg, size=10)
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return cell

    # Sort species by species-level score (highest rarity first)
    sorted_species = sorted(rarity_data.values(), key=lambda rd: (-rd.score, rd.species_key))

    row_i = 2
    for rd in sorted_species:
        # Get all size-class entries for this species, sorted by score descending
        sc_entries = sorted(
            [v for (sk, sb), v in size_class_rarity.items() if sk == rd.species_key],
            key=lambda v: (-v.score, v.size_bucket)
        )

        # If no size-class data, write a single row
        if not sc_entries:
            sc_entries = [None]

        for j, scr in enumerate(sc_entries):
            # Species-level columns only on first row of this species
            if j == 0:
                score_cell(ws, row_i, 1, rd.score, rd.score)
                ws.cell(row=row_i, column=2, value=rd.label).border = _border()
                ws.cell(row=row_i, column=2).font = Font(
                    bold=True, color=SCORE_COLORS.get(rd.score, ("000000",""))[0], size=10
                )
                ws.cell(row=row_i, column=3, value=rd.species_key).border = _border()
                ws.cell(row=row_i, column=3).font = Font(italic=True, size=10)
                new_flag = "🆕" if rd.new_to_system else ""
                ws.cell(row=row_i, column=12, value=new_flag).border = _border()
                if rd.new_to_system:
                    for c in range(1, 13):
                        ws.cell(row=row_i, column=c).fill = _fill("FFF59D")
            else:
                for c in (1, 2, 3, 12):
                    ws.cell(row=row_i, column=c).border = _border()

            # Size-class columns
            if scr:
                ws.cell(row=row_i, column=4, value=scr.size_label).border = _border()
                ws.cell(row=row_i, column=4).font = Font(size=10, color="616161")
                score_cell(ws, row_i, 5, scr.score, scr.score)
                ws.cell(row=row_i, column=6, value=scr.label).border = _border()
                ws.cell(row=row_i, column=6).font = Font(
                    color=SCORE_COLORS.get(scr.score, ("000000",""))[0], size=9
                )
                ws.cell(row=row_i, column=7, value=scr.vendor_count).border = _border()
                ws.cell(row=row_i, column=8, value=scr.obs_count).border = _border()
                ws.cell(row=row_i, column=9, value=len(rd.current_vendors) if j == 0 else "").border = _border()
                ws.cell(row=row_i, column=10, value=scr.last_seen or "").border = _border()
                lp_str = f"${scr.last_price:.2f}" if scr.last_price else ""
                ws.cell(row=row_i, column=11, value=lp_str).border = _border()

            row_i += 1

        # Light separator between species
        for c in range(1, 13):
            ws.cell(row=row_i - 1, column=c).border = Border(
                bottom=Side(style="thin", color="BDBDBD"),
                left=Side(style="thin"), right=Side(style="thin"),
            )
